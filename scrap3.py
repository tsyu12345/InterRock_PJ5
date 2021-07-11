from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, InvalidSwitchToTargetException, NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.common.keys import Keys
import openpyxl as px
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup as bs
import time
import datetime
import re
import requests as rq
import threading as th
import sys
import os
import PySimpleGUI as sig


class Job:
    """
    book_path = 'sample1-原本 - コピー.xlsx'
    book = px.load_workbook(book_path)
    sheet = book.worksheets[0]
    """
    book = px.Workbook()
    sheet = book.worksheets[0]
    RETRY = 3
    TIMEOUT = 15

    def __init__(self, path):
        self.save_path = path
        """
        dt_now = datetime.datetime.now()
        month = str(dt_now.month)
        day = str(dt_now.day)
        hour = str(dt_now.hour)
        min = str(dt_now.minute)
        self.name = month + day + "_result.xlsx"
        self.book_path = self.save_path + self.name
        """

    def init(self, driver_path, area, store_class):
        # init driver
        self.driver_path = resource_path(driver_path)
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("start-maximized")
        self.options.add_argument("enable-automation")
        self.options.add_argument("--headless")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-infobars")
        self.options.add_argument('--disable-extensions')
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-browser-side-navigation")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument('--ignore-certificate-errors')
        self.options.add_argument('--ignore-ssl-errors')
        prefs = {"profile.default_content_setting_values.notifications": 2}
        self.options.add_experimental_option("prefs", prefs)
        browser_path = resource_path(
            'Win_x64_857997_chrome-win\chrome-win\chrome.exe')
        self.options.binary_location = browser_path
        self.area_name = area
        self.store_class = store_class
        print(self.driver_path)
        print(browser_path)
        # init workBook
        """
        n_book = px.Workbook()
        n_sheet = n_book.worksheets[0]
        """
        menu = [
            "ジャンル",
            "店舗名",
            "店舗名カナ",
            "電話番号",
            "都道府県コード",
            "都道府県",
            "市区町村・番地",
            "店舗URL",
            "詳細データ取得日",
            "料金プラン着地",
            "月額料金",
            "お店のホームページ",
            "パンくず",
            "ヘッダー画像有無",
            "こだわり有無",
            "スライド画像数",
            "キャッチコピー",
            "アクセス・道案内",
            "営業時間",
            "定休日",
            "支払い方法",
            "設備",
            "カット価格",
            "席数",
            "スタッフ数",
            "駐車場",
            "こだわり条件",
            "備考",
            "スタッフ募集",
            "最終更新日"
        ]
        for c in range(1, 30+1):
            self.sheet.cell(row=1, column=c, value=menu[c-1])
        self.sheet.freeze_panes = "A2"
        """
        for r in self.sheet:
            for c in r:
                self.sheet[].font = font
        """
        """
        self.book_path = 'sample1-原本 - コピー.xlsx'
        self.book = px.load_workbook(self.book_path)
        self.sheet = self.book.worksheets[0]
        """

    def all_scrap(self):
        class_menu = [
            "ヘアサロン",
            "ネイル・まつげサロン",
            "リラクサロン",
            "エステサロン",
        ]
        for i in range(len(self.area_name)):
            for st_class in class_menu:
                self.store_class = st_class
                area = self.area_name[i]
                self.url_scrap(area)

    def clinick_url_scrap(self, area):
        url = 'https://clinic.beauty.hotpepper.jp/prefecture' + \
            str(self.call_jis_code(area))+'/'
        driver = webdriver.Chrome(
            executable_path=self.driver_path, options=self.options)
        driver.get(url=url)
        result_count = driver.find_element_by_css_selector(
            'body > div.l-content > div > div.l-two-columns > div.l-two-columns__item.l-two-columns__item--main > div.l-heading > p > span.c-search-result-heading__count').text
        result_pages = driver.find_element_by_css_selector(
            'body > div.l-content > div > div.l-two-columns > div.l-two-columns__item.l-two-columns__item--main > div.l-heading > p > span.c-search-result-heading__pagination > span.c-search-result-heading__max-page').text
        print(result_count)
        print(result_pages)
        for i in range(int(result_pages)):
            html = driver.page_source
            soup = bs(html, 'lxml')
            links_list = soup.select("p.clinic__name > a")
            for a in links_list:
                url = 'https://clinic.beauty.hotpepper.jp' + a.get('href')
                r = self.sheet.max_row
                self.sheet.cell(row=r+1, column=1, value=self.store_class)
                self.sheet.cell(row=r+1, column=8, value=url)
                print(self.sheet.cell(row=r+1, column=8).value)
            try:
                next_btn = driver.find_element_by_link_text("次へ")
                next_btn.click()
                time.sleep(1)
            except NoSuchElementException:
                break
        self.book_save()
        print("search complete")
        driver.quit()

    def clinick_info_scrap(self, start_index, end_index):
        driver = webdriver.Chrome(
            executable_path=self.driver_path, options=self.options)
        for i in range(start_index, end_index+1):
            driver.get(self.sheet.cell(row=i, column=8).value)
            html = driver.page_source
            soup = bs(html, 'lxml')
            cl_name = soup.select_one(
                'dt.clinic-header__clinic-name').get_text()
            cln_kana = soup.select_one(
                'dd.clinic-header__clinic-kana').get_text()
            catch_copy = soup.select_one(
                'p.clinic-overview__catchphrase').get_text()
            table_value = soup.select('div.table> table > tbody > tr > td')
            table_menu = soup.select('div.table > table > tbody > tr > th')
            for j, e in enumerate(table_menu):
                if e.get_text() == "住所":
                    all_address = table_value[j].get_text()
                    prefecture_search = re.search(
                        '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)
                    address_low = re.split(
                        '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)  # 県名とそれ以降を分離
                    prefecture = prefecture_search.group()  # 県名
                    jis_code = self.call_jis_code(prefecture)
                    municipality = address_low[1]  # それ以降
                    break
            print("都道府県：" + prefecture)
            print("市区町村番地：" + municipality)
            try:
                tel_tag = driver.find_element_by_link_text('番号を表示')
                tel_tag.click()
                tel_num = driver.find_element_by_xpath(
                    '/html/body/div[1]/div/div[2]/div[1]/div[2]/div[3]/div[2]/table/tbody/tr[1]/td/div/div/div/div/div[4]/p[2]').text
                print(tel_num)
            except NoSuchElementException:
                tel_num = ""
                pass
            try:
                credit = ""
                credit_s = soup.select(
                    'div.table> table > tbody > tr > td > ul.c-list > li.c-list__item')
                for i in range(len(credit_s)):
                    credit += credit_s[i].get_text() + "/"
            except:
                credit = ""

            slide_tag = soup.select('ol.clinic-overview__thumbnails > li')
            slide_cnt = len(slide_tag)

            self.sheet.cell(row=i, column=1, value=self.store_class)
            self.sheet.cell(row=i, column=2, value=cl_name)
            self.sheet.cell(row=i, column=3, value=cln_kana)
            self.sheet.cell(row=i, column=4, value=tel_num)
            self.sheet.cell(row=i, column=5, value=jis_code)
            self.sheet.cell(row=i, column=6, value=prefecture)
            self.sheet.cell(row=i, column=7, value=municipality)
            self.sheet.cell(row=i, column=9, value=self.scrap_day())
            self.sheet.cell(row=i, column=17, value=catch_copy)
            self.sheet.cell(row=i, column=16, value=slide_cnt)
            self.sheet.cell(row=i, column=21, value=credit)

    def url_scrap(self, area):
        #MAX_RETRY = 3
        print("starting ChromeDriver.exe....")
        driver = webdriver.Chrome(
            executable_path=self.driver_path, options=self.options)
        driver.get("https://beauty.hotpepper.jp/top/")  # top page
        sr_class = driver.find_element_by_link_text(self.store_class)
        sr_class.click()
        time.sleep(5)
        search = driver.find_element_by_css_selector('#freeWordSearch1')
        search.send_keys(area + Keys.ENTER)
        time.sleep(2)
        result_pages = driver.find_element_by_css_selector(
            'p.pa.bottom0.right0').text
        page_num = re.split('[/ ]', result_pages)
        pages = re.sub(r"\D", "", page_num[1])
        print("pages : " + pages)
        for i in range(int(pages)):
            try:
                html = driver.page_source
                soup = bs(html, 'lxml')
                links_list = soup.select("div.slcHeadContentsInner > h3 > a")
                for a in links_list:
                    url = a.get('href')
                    r = self.sheet.max_row
                    self.sheet.cell(row=r+1, column=1, value=self.store_class)
                    self.sheet.cell(row=r+1, column=6, value=area)
                    self.sheet.cell(row=r+1, column=8, value=url)
                    print(self.sheet.cell(row=r+1, column=8).value)
                try:
                    pre_url = driver.current_url
                    pre_index = i
                    next_btn = driver.find_element_by_link_text("次へ")
                    next_btn.click()
                    time.sleep(1)
                except NoSuchElementException:
                    break
            except (WebDriverException, TimeoutException):
                self.book_save()
                driver.quit()
                time.sleep(10)
                print("starting ChromeDriver.exe....")
                driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
                driver.get(pre_url)
                next_btn = driver.find_element_by_link_text("次へ")
                next_btn.click()
                time.sleep(1)
                continue
            else:
                pass
        self.book_save()
        print("search complete")
        driver.quit()

    def info_scrap(self, start_index, end_index):
        counter = 0
        driver = webdriver.Chrome(
            executable_path=self.driver_path, options=self.options)
        for i in range(start_index, end_index+1):
            counter += 1
            try:
                if counter % 100 == 0:
                    driver.quit()
                    time.sleep(3)
                    driver = webdriver.Chrome(
                        executable_path=self.driver_path, options=self.options)
                driver.get(self.sheet.cell(row=i, column=8).value)
            except (WebDriverException, TimeoutException):
                print('retry....')
                for k in range(3):
                    try:
                        time.sleep(10)
                        print('restart browser...')
                        driver.quit()
                        driver = webdriver.Chrome(
                            executable_path=self.driver_path, options=self.options)
                        print('open browser...')
                        time.sleep(3)
                        driver.get(self.sheet.cell(row=i, column=8).value)
                    except:
                        print('Error not Resolved')
                        continue
            else:
                pass
            html = driver.page_source
            soup = bs(html, 'lxml')
            table_value = soup.select(
                'div.mT30 > table > tbody > tr > td')
            table_menu = soup.select(
                'div.mT30 > table > tbody > tr > th')
            print(table_menu)
            print(table_value)
            # 住所の抽出（少々処理があるため別で書き出す）
            pref_tf = True
            for j, e in enumerate(table_menu):
                if e.get_text() == "住所":
                    all_address = table_value[j].get_text()
                    prefecture_search = re.search(
                        '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)
                    address_low = re.split(
                        '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)  # 県名とそれ以降を分離
                    prefecture = prefecture_search.group()  # 県名
                    if prefecture != self.sheet.cell(row=i, column=6).value:
                        self.sheet.cell(row=i, column=8, value="")
                        self.sheet.cell(row=i, column=6, value="")
                        self.sheet.cell(row=i, column=1, value="")
                        pref_tf = False
                    jis_code = self.call_jis_code(prefecture)
                    municipality = address_low[1]  # それ以降
                    break
            print("都道府県：" + prefecture)
            print("市区町村番地：" + municipality)
            # 指定エリアでないとき、下記処理を行わない
            try:
                if pref_tf:
                    store_name_tag = soup.select_one('p.detailTitle > a')
                    store_name = store_name_tag.get_text()
                    print("店名：" + store_name)
                    st_name_kana_tag = soup.select_one(
                        'div > p.fs10.fgGray')
                    st_name_kana = st_name_kana_tag.get_text()
                    print("店名カナ：" + st_name_kana)
                    try:
                        tel_tag = soup.select_one(
                            'div.mT30 > table > tbody > tr > td > a')
                        tel_url = tel_tag.get('href')
                        respons_tel = rq.get(tel_url)
                        html_tel = respons_tel.text
                        soup_tel = bs(html_tel, 'lxml')
                        tel_num_tag = soup_tel.select_one(
                            'table > tr > td')
                        tel_num = tel_num_tag.get_text()
                        tel_num = str(tel_num)
                        tel_num = tel_num.replace(' ', "")
                        print("TEL : " + tel_num)
                    except:
                        tel_num = ""
                        pass
                        # ヘッダー画像の有無
                    try:
                        driver.find_element_by_css_selector(
                            'div.slnHeaderSliderPhoto.jscViewerPhoto')
                        head_img_yn = "有"
                        # self.sheet.cell(row=index, column=14, value="有")
                    except NoSuchElementException:
                        head_img_yn = "無"
                        # self.sheet.cell(row=index, column=14, value="無")
                        pass
                    catch_copy_tag = soup.select_one('div > p > b > strong')
                    catch_copy = catch_copy_tag.get_text()

                    pankuzu_tag = soup.select('#preContents > ol > li')
                    pankuzu = ""
                    for pan in pankuzu_tag:
                        pankuzu += pan.get_text()
                    print(pankuzu)

                    slide_img_tag = soup.select(
                        'div.slnTopImgCarouselWrap.jscThumbWrap > ul > li')
                    slide_cnt = len(slide_img_tag)

                    # write Excel
                    self.sheet.cell(row=i, column=2, value=store_name)
                    self.sheet.cell(row=i, column=3, value=st_name_kana)
                    self.sheet.cell(row=i, column=4, value=tel_num)
                    self.sheet.cell(row=i, column=5, value=jis_code)
                    self.sheet.cell(row=i, column=6, value=prefecture)
                    self.sheet.cell(row=i, column=7, value=municipality)
                    self.sheet.cell(row=i, column=9, value=self.scrap_day())
                    self.sheet.cell(row=i, column=13, value=pankuzu)
                    self.sheet.cell(row=i, column=16, value=slide_cnt)
                    self.sheet.cell(row=i, column=17, value=catch_copy)
                    self.sheet.cell(row=i, column=14, value=head_img_yn)
                    # 他の情報処理
                    try:
                        for j in range(2, len(table_value)):
                            for c in range(1, self.sheet.max_column):
                                if table_menu[j].get_text() == self.sheet.cell(row=1, column=c).value:
                                    self.sheet.cell(
                                        row=i, column=c, value=table_value[j].get_text())
                                    break
                    except RuntimeError:
                        pass
                else:
                    print("prefname is False")
            except:
                self.book_save()
                pass
        driver.quit()
        # self.book_save()

    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": '01',
            "青森県": '02',
            "岩手県": '03',
            "宮城県": '04',
            "秋田県": '05',
            "山形県": '06',
            "福島県": '07',
            "茨城県": '08',
            "栃木県": '09',
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        print(code)
        return code

    def scrap_day(self):
        dt_now = datetime.datetime.now()
        year = str(dt_now.year)
        month = str(dt_now.month)
        day = str(dt_now.day)
        hour = str(dt_now.hour)
        min = str(dt_now.minute)
        data_day = year + "," + month + day + "," + hour + min
        print(data_day)
        return data_day

    def book_save(self):
        self.book.save(self.save_path)


# other functions↓
"""
def progress_per(percent):
    PopupWindow().prg_per += percent
"""


def apper_adjst(path):
    book = px.load_workbook(path)
    sheet = book.worksheets[0]
    for r in range(2, sheet.max_row+1):
        print("check" + str(r))
        if sheet.cell(row=r, column=1).value == "" or sheet.cell(row=r, column=1).value == None:
            sheet.delete_rows(r)
    book.save(path)


def retry_info_scrap(path):
    job = Job(path)
    driver_path = resource_path('chromedriver_win32/chromedriver.exe')
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-gpu')
    browser_path = resource_path(
        'Win_x64_857997_chrome-win\chrome-win\chrome.exe')
    options.binary_location = browser_path
    driver = webdriver.Chrome(executable_path=driver_path, options=options)
    book = px.load_workbook(path)
    sheet = book.worksheets[0]
    counter = 0
    for i in range(2, sheet.max_row+1):
        print('check : ' + str(i))
        if sheet.cell(row=i, column=2).value in (" ", "", None):
            counter += 1
            try:
                if counter % 100 == 0:
                    driver.quit()
                    time.sleep(3)
                    driver = webdriver.Chrome(
                        executable_path=driver_path, options=options)
                driver.get(sheet.cell(row=i, column=8).value)
            except (WebDriverException, TimeoutException):
                print('retry....')
                for k in range(3):
                    try:
                        time.sleep(10)
                        print('restart browser...')
                        driver.quit()
                        driver = webdriver.Chrome(
                            executable_path=driver_path, options=options)
                        print('open browser...')
                        time.sleep(3)
                        driver.get(sheet.cell(row=i, column=8).value)
                    except:
                        print('Error not Resolved')
                        continue
            else:
                pass
            try:
                html = driver.page_source
                soup = bs(html, 'lxml')
                table_value = soup.select(
                    'div.mT30 > table > tbody > tr > td')
                table_menu = soup.select(
                    'div.mT30 > table > tbody > tr > th')
                print(table_menu)
                print(table_value)
                # 住所の抽出（少々処理があるため別で書き出す）
                pref_tf = True
                for j, e in enumerate(table_menu):
                    if e.get_text() == "住所":
                        all_address = table_value[j].get_text()
                        prefecture_search = re.search(
                            '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)
                        address_low = re.split(
                            '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)  # 県名とそれ以降を分離
                        prefecture = prefecture_search.group()  # 県名
                        if prefecture != sheet.cell(row=i, column=6).value:
                            sheet.cell(row=i, column=8, value="")
                            sheet.cell(row=i, column=6, value="")
                            sheet.cell(row=i, column=1, value="")
                            pref_tf = False
                        jis_code = job.call_jis_code(prefecture)
                        municipality = address_low[1]  # それ以降
                        break
                print("都道府県：" + prefecture)
                print("市区町村番地：" + municipality)
                # 指定エリアでないとき、下記処理を行わない
                if pref_tf:
                    store_name_tag = soup.select_one('p.detailTitle > a')
                    store_name = store_name_tag.get_text()
                    print("店名：" + store_name)
                    st_name_kana_tag = soup.select_one(
                        'div > p.fs10.fgGray')
                    st_name_kana = st_name_kana_tag.get_text()
                    print("店名カナ：" + st_name_kana)
                    try:
                        tel_tag = soup.select_one(
                            'div.mT30 > table > tbody > tr > td > a')
                        tel_url = tel_tag.get('href')
                        respons_tel = rq.get(tel_url)
                        html_tel = respons_tel.text
                        soup_tel = bs(html_tel, 'lxml')
                        tel_num_tag = soup_tel.select_one(
                            'table > tr > td')
                        tel_num = tel_num_tag.get_text()
                        print("TEL : " + tel_num)
                    except:
                        tel_num = ""
                        pass
                        # ヘッダー画像の有無
                    try:
                        driver.find_element_by_css_selector(
                            'div.slnHeaderSliderPhoto.jscViewerPhoto')
                        head_img_yn = "有"
                        # self.sheet.cell(row=index, column=14, value="有")
                    except NoSuchElementException:
                        head_img_yn = "無"
                        # self.sheet.cell(row=index, column=14, value="無")
                        pass
                    catch_copy_tag = soup.select_one('div > p > b > strong')
                    catch_copy = catch_copy_tag.get_text()

                    pankuzu_tag = soup.select('#preContents > ol > li')
                    pankuzu = ""
                    for pan in pankuzu_tag:
                        pankuzu += pan.get_text()
                    print(pankuzu)

                    slide_img_tag = soup.select(
                        'div.slnTopImgCarouselWrap.jscThumbWrap > ul > li')
                    slide_cnt = len(slide_img_tag)

                    # write Excel
                    sheet.cell(row=i, column=2, value=store_name)
                    sheet.cell(row=i, column=3, value=st_name_kana)
                    sheet.cell(row=i, column=4, value=tel_num)
                    sheet.cell(row=i, column=5, value=jis_code)
                    sheet.cell(row=i, column=6, value=prefecture)
                    sheet.cell(row=i, column=7, value=municipality)
                    sheet.cell(row=i, column=9, value=job.scrap_day())
                    sheet.cell(row=i, column=13, value=pankuzu)
                    sheet.cell(row=i, column=16, value=slide_cnt)
                    sheet.cell(row=i, column=17, value=catch_copy)
                    sheet.cell(row=i, column=14, value=head_img_yn)
                    # 他の情報処理
                    try:
                        for j in range(2, len(table_value)):
                            for c in range(1, sheet.max_column):
                                if table_menu[j].get_text() == sheet.cell(row=1, column=c).value:
                                    sheet.cell(
                                        row=i, column=c, value=table_value[j].get_text())
                                    break
                    except RuntimeError:
                        pass
                else:
                    print("prefname is False")
                # book.save(path)
            except:
                book.save(path)
                pass
    book.save(path)
    driver.quit()
    # self.book_save()


def check(path):
    book = px.load_workbook(path)
    sheet = book.worksheets[0]
    for r in range(2, sheet.max_row+1):
        print("check" + str(r))
        if sheet.cell(row=r, column=2).value in ("", " ", None):
            fill =  PatternFill(patternType='solid', fgColor='ffff00')
            sheet['B'+str(r)].fill = fill
            sheet.cell(row=r, column=2, value='抽出不可')
        if sheet.cell(row=r, column=1).value == "" or sheet.cell(row=r, column=1).value == None:
            return False
    return True


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


def range_segmentation(maxindex):
    rang = list(range(2, maxindex, int(maxindex / 4)))
    print(rang)
    return rang


def process0(path, area, st_class):
    job = Job(path)
    # progress_per(10)
    job.init('chromedriver_win32\chromedriver.exe', area, st_class)
    if st_class == 'すべてのジャンル':
        job.all_scrap()
    # progress_per(10)
    else:
        for i in range(len(area)):
            job.url_scrap(area=area[i])
    print('URL scrap complete.')
    job.book_save()


def process1(s_index, e_index, path, area, st_class):
    job = Job(path)
    job.init('chromedriver_win32\chromedriver.exe', area, st_class)
    job.info_scrap(s_index, e_index)
    # job.book.save()


def load_max_row(path):
    job = Job(path)
    max_row = job.sheet.max_row
    return max_row


def multiThread(path, area, st_class):
    job = Job(path)

    # progress_per(10)
    sheet_index = load_max_row(path)
    rng = range_segmentation(sheet_index)
    th_s = []
    for i in range(3):
        if i == 0:
            add = th.Thread(target=process1, args=(
                rng[0], rng[1], path, area, st_class))
        else:
            add = th.Thread(target=process1, args=(
                rng[i]+1, rng[i+1], path, area, st_class))
        th_s.append(add)
    add = th.Thread(target=process1, args=(
        rng[3]+1, sheet_index, path, area, st_class))
    th_s.append(add)

    for i in range(len(th_s)):
        # progress_per(2)
        th_s[i].start()
    for i in range(len(th_s)):
        th_s[i].join()
    job.book_save()


def main(path, area, st_class):
    try:
        s_time = time.time()
        print("scrap start...")
        process0(path, area, st_class)
        multiThread(path, area, st_class)
        print("scrap end.")
        while check(path) == False:
            apper_adjst(path)
            print('check False')
        print('check True')
        print('取得できなかった情報がないか検査中…')
        retry_info_scrap(path)
        print('終了')
        while check(path) == False:
            apper_adjst(path)
            print('check False')
        print('check True')
        sig.popup('お疲れ様でした。抽出終了です。ファイルを確認してください。\n保存先：' +
                  path, keep_on_top=True)
        e_time = time.time() - s_time
        p_time = "{0}".format(e_time) + "[sec]"
        # progress_per(10)
    except RuntimeError:
        pass


"""
area_name = gui.value['pref_name']
st_class  = gui.value['store_class']
save_path = gui.value['path']
process0(save_path, area_name, st_class)
multiThread(save_path, area_name, st_class)
"""
if __name__ == "__main__":
    # main('logs/result.xlsx', ['高知県'], 'ネイル・まつげサロン')
    path = "C:/Users/syu/Documents/Tokyo.xlsx"
    while check(path) == False:
        apper_adjst(path)
        print('check False')
    print('check True')
    print('取得できなかった情報がないか検査中…')
    retry_info_scrap(path)
    print('終了')
    while check(path) == False:
        apper_adjst(path)
        print('check False')
    print('check True')
