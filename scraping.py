#from multiprocessing.pool import ApplyResult
#import queue
#from typing import Tuple
from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, InvalidSessionIdException, InvalidSwitchToTargetException, NoSuchElementException, TimeoutException, WebDriverException
#from selenium.webdriver.chrome.webdriver import WebDriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl as px
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup as bs
from multiprocessing import Pool, Manager
from urllib3.exceptions import MaxRetryError
#from concurrent.futures import ProcessPoolExecutor
#import threading as th
import time
import datetime
import re
import requests as rq
import sys
import os
import subprocess

class ScrapingURL(object):
    def __init__(self, path, row_counter, sync_data_list):
        self.path = path
        self.driver_path = 'chromedriver_win32/chromedriver.exe'
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("start-maximized")
        self.options.add_argument("enable-automation")
        self.options.add_argument("--headless")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-infobars")
        self.options.add_argument('--disable-extensions')
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-browser-side-navigation")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument('--ignore-certificate-errors')
        self.options.add_argument('--ignore-ssl-errors')
        prefs = {"profile.default_content_setting_values.notifications": 2}
        self.options.add_experimental_option("prefs", prefs)
        browser_path = 'chrome-win/chrome.exe'
        self.options.binary_location = browser_path
        #driver = webdriver.Chrome(executable_path=driver_path, options=self.options)
        self.writeRow = 0
        self.row_counter = row_counter
        self.url_list = sync_data_list
        self.table_list = [
        ]

    def all_scrap(self, area_list):  # 全件抽出
        class_menu = [
            "ヘアサロン",
            "ネイル・まつげサロン",
            "リラクサロン",
            "エステサロン",
        ]
        for area in area_list:
            for junle in class_menu:
                self.__url_scrap(area, junle)
            print(area + "search complete.")

    def search(self, area_list, store_junle):
        for area in area_list:
            self.__url_scrap(area, store_junle)

    def __url_scrap(self, area, store_junle):
        #MAX_RETRY = 3
        self.sub_driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        print("starting ChromeDriver.exe....")
        wait = WebDriverWait(self.sub_driver, 180)#Max wait time(second):180s
        self.sub_driver.get("https://beauty.hotpepper.jp/top/")  # top page
        sr_class = self.sub_driver.find_element_by_link_text(store_junle)#ジャンル選択
        sr_class.click()
        wait.until(EC.visibility_of_element_located((By.ID, "freeWordSearch1")))
        search = self.sub_driver.find_element_by_id('freeWordSearch1')
        search.send_keys(area + Keys.ENTER)
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "p.pa.bottom0.right0")))
        result_pages = self.sub_driver.find_element_by_css_selector('p.pa.bottom0.right0').text
        page_num = re.split('[/ ]', result_pages)
        pages = re.sub(r"\D", "", page_num[1])
        print("pages : " + pages)
        self.page_count = int(pages)
        for i in range(int(pages)):
            if i % 100 == 0 and i > 0:
                cur_url = self.sub_driver.current_url
                self.sub_driver.quit()
                time.sleep(5)
                self.sub_driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
                self.sub_driver.get(cur_url)
                
            #sig.OneLineProgressMeter("掲載URLの抽出中...", self.counter, int(pages))
            try:
                html = self.sub_driver.page_source
                soup = bs(html, 'lxml')
                if store_junle == 'ヘアサロン':
                    links_list = soup.select("#mainContents > ul > li > div.slnCassetteHeader > h3 > a")
                else:
                    links_list = soup.select("#mainContents > ul > li > div.slcHeadWrap > div > div.slcHeadContentsInner > h3 > a")
                #print(links_list)
                #sheet.max_rowの呼び出しの際に発生するRunTimeError対策
                
                for a in links_list:
                    url = a.get('href')
                    print(url)
                    self.writeRow += 1
                    add = [store_junle, area, url]
                    self.url_list.append(add)
                    self.row_counter.value += 1
                    """
                    self.sheet.cell(row=write_row, column=1, value=store_junle)#ジャンル
                    self.sheet.cell(row=write_row, column=6, value=area)#エリア
                    self.sheet.cell(row=write_row, column=8, value=url)#URL
                    """
                    #print(self.sheet.cell(row=r+1, column=8).value)
                #print(self.sheet.max_row)
                print(self.row_counter)
                #self.book.save(self.path)
                #Issue:↑で一時保存するとinfo_scrap()との衝突するのかPermissionErrorが発生する場合がある。
                try:
                    pre_url = self.sub_driver.current_url
                    #pre_index = i
                    next_btn = self.sub_driver.find_element_by_link_text("次へ")
                    next_btn.click()
                    wait.until(EC.visibility_of_all_elements_located)
                except NoSuchElementException:
                    break
            except (WebDriverException, TimeoutException):
                #self.book.save(self.path)
                self.sub_driver.quit()
                time.sleep(10)
                print("starting ChromeDriver.exe....")
                self.sub_driver = webdriver.Chrome(
                    executable_path=self.driver_path, options=self.options)
                self.sub_driver.get(pre_url)
                next_btn = self.sub_driver.find_element_by_link_text("次へ")
                next_btn.click()
                wait.until(EC.visibility_of_all_elements_located)
                continue
            else:
                pass
        
        #self.book.save(self.path)
        print("search complete")
        self.sub_driver.quit()
        #driver.close()

class ScrapingInfomation(ScrapingURL):
    def __init__(self, path, row_counter, url_list_data, end_count, info_datas):
        """
        path : WorkSheetPath \n
        row_counter : Manager.Value('i', 0) \n
        url_list_data : Manager.list() \n
        get_new_driver : Manager.Value('b') ※共有メモリ上のboolean変数 \n

        """
        super(ScrapingInfomation, self).__init__(path, row_counter, url_list_data)
        #driver = webdriver.Chrome(executable_path=driver_path, options=self.options)
        #self.thread_local = th.local()
        self.end_count = end_count
        self.info_datas = info_datas #結果格納用リスト
        self.table_menu = {
            12:'お店のホームページ',
            18:'アクセス・道案内',
            19:'営業時間',
            20:'定休日',
            21:'支払い方法',
            22:'設備',
            23:'カット価格',
            24:'席数',
            25:'スタッフ数',
            26:'駐車場',
            27:'こだわり条件',
            28:'備考',
            29:'スタッフ募集',
        }
    
    def loadHtml(self, store_url_data:list,):
        #conunter  
        driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        wait = WebDriverWait(driver, 180)
        load_counter = 0
        for url_data in store_url_data:
            
            if load_counter % 100 == 0 and load_counter != 0: #メモリ対策でブラウザを再起動。
                driver.delete_all_cookies()
                driver.quit()
                time.sleep(10)
                driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
                wait = WebDriverWait(driver, 180)
            try:
                url = url_data[2] #[[junle, area, url], [],....]の想定
                #MAXRetryError↓ load_counter = 2049
                driver.get(url)
            except (WebDriverException, TimeoutException, MaxRetryError):
                driver.delete_all_cookies()
                driver.quit()
                time.sleep(30)#強制30秒待機
                wait = WebDriverWait(driver, 180)
                driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
                driver.get(url)
                #issues: urlに''が渡されるとInvaild argument Exceptionが発生し処理が止まる。
            else:
                pass
            wait.until(EC.visibility_of_all_elements_located)
            html = driver.page_source
            data_list:list = self.__extraction(html, url_data)
            if data_list != []:
                self.info_datas.append(data_list)
            else:
                pass
            self.end_count.value += 1
            load_counter += 1
        driver.quit()
        

    def __create_data_list(self, data_length):
        """
        指定数分のNone要素のみのリストを作成。\n
        data = [None, None, None ......(len() = data_length)]
        """
        data_list = []
        for i in range(data_length):
            data_list.append(None)
        return data_list

    def __extraction(self, html, store_url_data):
        """
        HTMLを受け取り、情報を抽出,結果のリストを返す。
        """
        data_list = self.__create_data_list(31)
        soup = bs(html, 'lxml')
        table_value = soup.select(
            'div.mT30 > table > tbody > tr > td')
        table_menu = soup.select(
            'div.mT30 > table > tbody > tr > th')
          # 住所の抽出（少々処理があるため別で書き出す）
        pref_tf = True
        prefecture = ""
        for j, e in enumerate(table_menu):
            if e.get_text() == "住所":
                all_address = table_value[j].get_text()
                prefecture_search = re.search(
                    '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)
                address_low = re.split(
                    '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)  # 県名とそれ以降を分離
                prefecture = prefecture_search.group()  # 県名
                print(prefecture)
                if prefecture != store_url_data[1]:
                    #self.url_list.remove(store_url_data)
                    return []
                    """
                    self.sheet.cell(row=index, column=8, value='')
                    self.sheet.cell(row=index, column=6, value='')
                    self.sheet.cell(row=index, column=1, value='')
                    """
                jis_code = self.call_jis_code(prefecture)
                municipality = address_low[1]  # それ以降
                break
        # 指定エリアでないとき、下記処理を行わない
        try:
            store_name_tag = soup.select_one('p.detailTitle > a')
            store_name = store_name_tag.get_text() if store_name_tag != None else None
            #store_name = store_name_tag.get_text()
            print("店名：" + store_name)
            st_name_kana_tag = soup.select_one(
                'div > p.fs10.fgGray')
            st_name_kana = st_name_kana_tag.get_text() if st_name_kana_tag != None else None
            #st_name_kana = st_name_kana_tag.get_text()
            print("店名カナ：" + st_name_kana)
            try:
                tel_tag = soup.select_one(
                    'div.mT30 > table > tbody > tr > td > a')
                tel_url = tel_tag.get('href')
                respons_tel = rq.get(tel_url)
                html_tel = respons_tel.text
                soup_tel = bs(html_tel, 'lxml')
                tel_num_tag = soup_tel.select_one(
                    'table > tr > td')
                tel_num = tel_num_tag.get_text() if tel_num_tag != None else None
                #tel_num = tel_num_tag.get_text()
                tel_num = str(tel_num)
                tel_num = tel_num.replace(' ', "")
                print("TEL : " + tel_num)
            except:
                tel_num = None
                pass
                # ヘッダー画像の有無
            head_img_tag = soup.select_one('div.slnHeaderSliderPhoto.jscViewerPhoto')    
            head_img_yn = "有" if head_img_tag != None else "無"

            catch_copy_tag = soup.select_one('div > p > b > strong')
            catch_copy = catch_copy_tag.get_text() if catch_copy_tag != None else None
            #catch_copy = catch_copy_tag.get_text()
            pankuzu_tag = soup.select('#preContents > ol > li')
            pankuzu = ""
            for pan in pankuzu_tag:
                pankuzu += pan.get_text() if pan != None else ""
            print(pankuzu)

            slide_img_tag = soup.select(
                'div.slnTopImgCarouselWrap.jscThumbWrap > ul > li')
            slide_cnt = len(slide_img_tag)
        
        #append data_list
            data_list[0] = store_url_data[0]
            data_list[1] = store_name
            data_list[2] = st_name_kana
            data_list[3] = tel_num
            data_list[4] = jis_code
            data_list[5] = prefecture
            data_list[6] = municipality
            data_list[7] = store_url_data[2]
            data_list[8] = self.__scrap_day()
            data_list[12] = pankuzu
            data_list[15] = slide_cnt
            data_list[16] = catch_copy
            data_list[13] = head_img_yn 

            for j in range(2, len(table_value)):
                for row, menu in zip(self.table_menu.keys(), self.table_menu.values()):  
                    if table_menu[j].get_text() == menu:
                        data_list[row-1] = table_value[j].get_text()
                        break
        except:
            pass
        return data_list

    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": '01',
            "青森県": '02',
            "岩手県": '03',
            "宮城県": '04',
            "秋田県": '05',
            "山形県": '06',
            "福島県": '07',
            "茨城県": '08',
            "栃木県": '09',
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        print(code)
        return code

    def __scrap_day(self):
        dt_now = datetime.datetime.now()
        year = str(dt_now.year)
        month = str(dt_now.month)
        day = str(dt_now.day)
        hour = str(dt_now.hour)
        min = str(dt_now.minute)
        data_day = year + "," + month + day + "," + hour + min
        print(data_day)
        return data_day


class WriteWorkBook():
    book = px.Workbook()
    sheet = book.worksheets[0]
    def __init__(self, path, end_count):
        self.path = path
        self.end_count = end_count #共有メモリ上のカウンタ変数
    
    def init_work_book(self):
        menu = [
            "ジャンル",
            "店舗名",
            "店舗名カナ",
            "電話番号",
            "都道府県コード",
            "都道府県",
            "市区町村・番地",
            "店舗URL",
            "詳細データ取得日",
            "料金プラン着地",
            "月額料金",
            "お店のホームページ",
            "パンくず",
            "ヘッダー画像有無",
            "こだわり有無",
            "スライド画像数",
            "キャッチコピー",
            "アクセス・道案内",
            "営業時間",
            "定休日",
            "支払い方法",
            "設備",
            "カット価格",
            "席数",
            "スタッフ数",
            "駐車場",
            "こだわり条件",
            "備考",
            "スタッフ募集",
            "最終更新日"
        ]
        for c in range(1, 30+1):
            self.sheet.cell(row=1, column=c, value=menu[c-1])
        self.sheet.freeze_panes = "A2"
        self.book.save(self.path)
    
    def wirte_data(self, data_list:list, index:int):
        """
        1プロセスあたりのスクレイピングデータを書き込む。\n
        data_list = [junle, store_name, st_name_kana, ....]
        """
        for col in range(1, 30+1):
            try:
                self.sheet.cell(row=index, column=col, value=data_list[col-1])
            except IndexError:
                return False
            #self.end_count.value += 1
        return True 
        



class Implementation():
    def __init__(self, path, area_list, junle):
        self.area_list = area_list
        self.junle = junle
        self.manager = Manager() #共有メモリ
        self.max_row_counter = self.manager.Value('i', 0) #最大読み込み行数格納用
        self.scrap_url_list = self.manager.list() #URL格納用リスト
        self.info_datas = self.manager.list() #結果格納用リスト
        self.end_count = self.manager.Value('i', 0)
        self.search = ScrapingURL(path, self.max_row_counter, self.scrap_url_list)
        self.scrap = ScrapingInfomation(path, self.max_row_counter, self.scrap_url_list, self.end_count, self.info_datas)
        self.scrap2 = ScrapingInfomation(path, self.max_row_counter, self.scrap_url_list, self.end_count, self.info_datas)
        self.writeBook = WriteWorkBook(path, self.end_count,)
        self.writeBook.init_work_book()
        #driver1 = webdriver.Chrome(executable_path=self.scrap.driver_path, options=self.scrap.options)
        #driver2 = webdriver.Chrome(executable_path=self.scrap.driver_path, options=self.scrap.options)
        #driver_list = self.manager.list()
        self.search_sum = 1
        self.writed_index = 0 #結果格納リストの書き込み済みindex数
        self.book_index = 2 #ワークシートへの書き込み済み行数。
        self.list1 = []#リクエストURLの格納用
        self.list2 = []
        self.scraped_url_index = 0 #リクエスト済みのURLリストのindex数
        self.p = Pool()
        self.search_flg = False

    def call_prog_value(self):
        return self.end_count.value, self.search_sum

    def run(self):
    #row_counter = Value('i', 0)
    #manager = Manager()
    #max_row_counter = manager.Value('i', 0)
    #scrap_url_list = manager.list()
    #executer = ProcessPoolExecutor(max_workers=None)
    #futures = executer.submit(test.url_scrap, '高知県', 'ヘアサロン')
        if self.junle == 'すべてのジャンル':
            future = self.p.apply_async(self.search.all_scrap, args=([self.area_list]))
        else:
            future = self.p.apply_async(self.search.search, args=([self.area_list, self.junle]))

        scrap_flg = True
        self.search_flg = True
    
        while scrap_flg:
            self.search_sum = self.search.row_counter.value#表示更新用
            #２つのdriverで並列にスクレイピングするため、共有メモリのリストから個別に分けてdataを読み込む。
            if len(self.scrap_url_list) > 0:
                self.create_url_data_list()
                result1 = self.p.apply_async(self.scrap2.loadHtml, args=([self.list1,]))
                result2 = self.p.apply_async(self.scrap.loadHtml, args=([self.list2,]))

                #join scraping and Writing work book.
                doing = True
                async_result = [False, False]
                while doing:
                    self.search_sum = self.search.row_counter.value
                    if False not in async_result:
                        doing = False
                        self.info_datas_writing()
                        break
                    if result1.ready():
                        print("result1 end")
                        async_result[0] = True
                        self.list1.clear()
                        
                    if result2.ready():
                        print("result2 end")
                        async_result[1] = True
                        self.list2.clear()        

                self.search_sum = self.search.row_counter.value

            if future.ready():
                future.get()
                self.search_sum = self.search.row_counter.value
                self.search_flg = False

            if (self.search_flg == False and 
                scrap_flg == True and 
                (len(self.scrap_url_list) == 0 and len(self.list1) == 0 and len(self.list2)== 0)):
                scrap_flg = False
                self.info_datas_writing()
                print("break!!")
                self.writeBook.book.save(self.writeBook.path)
                break

    def create_url_data_list(self):
        addDataLength = len(self.scrap_url_list) #呼び出し時点での大きさを取得
        print("now scrap end index :" + str(self.scraped_url_index))
        for i in range(0, addDataLength): #その時点での大きさまで各配列へ格納する。
            if i % 2 == 0:
                self.list1.append(self.scrap_url_list.pop(0))
            else:
                self.list2.append(self.scrap_url_list.pop(0))
       
    def cancel(self):
        try:
            self.info_datas_writing()
            self.writeBook.book.save(self.writeBook.path)
            self.p.terminate()
        except Exception:
            pass

    def info_datas_writing(self):
        """
        共有メモリ上のリストを監視し、キューがemptyでない限り書き込みを続ける。
        """
        print("called!")
        length = len(self.info_datas)
        #print("write_data_len : " + str(length))
        for i in range(0, length):
            #print(self.info_datas.empty())
            #print(self.info_datas[0])
            #data = self.info_datas[i]
            #print(data)
            self.writeBook.wirte_data(self.info_datas.pop(0), self.book_index)
                    #self.end_count += 1
            #self.writed_index += 1
            self.book_index += 1
        self.writeBook.book.save(self.writeBook.path)


def resource_path(relative_path):#バイナリフィルのパスを提供
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


    
if __name__ == '__main__':
    test = Implementation('concurrent-test.xlsx', ['高知県','徳島県'], 'ヘアサロン')
    test.run()

