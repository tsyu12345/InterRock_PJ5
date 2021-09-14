from selenium import webdriver
from selenium.common.exceptions import InvalidArgumentException, InvalidSwitchToTargetException, NoSuchElementException, TimeoutException, WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl as px
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup as bs
from multiprocessing import Pool 
import threading as th
import time
import datetime
import re
import requests as rq
import sys
import os

class Scraping():
    book = px.Workbook()
    sheet = book.worksheets[0]
    RETRY = 3
    TIMEOUT = 15

    def __init__(self, path):
        self.path = path
        # init driver
        self.driver_path = resource_path('chromedriver_win32/chromedriver.exe')
        self.options = webdriver.ChromeOptions()
        self.options.add_argument("start-maximized")
        self.options.add_argument("enable-automation")
        #self.options.add_argument("--headless")
        self.options.add_argument("--no-sandbox")
        self.options.add_argument("--disable-infobars")
        self.options.add_argument('--disable-extensions')
        self.options.add_argument("--disable-dev-shm-usage")
        self.options.add_argument("--disable-browser-side-navigation")
        self.options.add_argument("--disable-gpu")
        self.options.add_argument('--ignore-certificate-errors')
        self.options.add_argument('--ignore-ssl-errors')
        prefs = {"profile.default_content_setting_values.notifications": 2}
        self.options.add_experimental_option("prefs", prefs)
        browser_path = resource_path('chrome-win/chrome.exe')
        self.options.binary_location = browser_path
        self.driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        self.counter = 0
        self.page_count = 1
        #self.search_end_flg = False
        #self.area = area
        #self.store_class = store_class

    def init_work_book(self):
        menu = [
            "ジャンル",
            "店舗名",
            "店舗名カナ",
            "電話番号",
            "都道府県コード",
            "都道府県",
            "市区町村・番地",
            "店舗URL",
            "詳細データ取得日",
            "料金プラン着地",
            "月額料金",
            "お店のホームページ",
            "パンくず",
            "ヘッダー画像有無",
            "こだわり有無",
            "スライド画像数",
            "キャッチコピー",
            "アクセス・道案内",
            "営業時間",
            "定休日",
            "支払い方法",
            "設備",
            "カット価格",
            "席数",
            "スタッフ数",
            "駐車場",
            "こだわり条件",
            "備考",
            "スタッフ募集",
            "最終更新日"
        ]
        for c in range(1, 30+1):
            self.sheet.cell(row=1, column=c, value=menu[c-1])
        self.sheet.freeze_panes = "A2"

        self.book.save(self.path)

    def all_scrap(self, area_list):  # 全件抽出
        class_menu = [
            "ヘアサロン",
            "ネイル・まつげサロン",
            "リラクサロン",
            "エステサロン",
        ]
        for area in area_list:
            for junle in class_menu:
                self.counter = 0
                self.url_scrap(area, junle)
            print(area + "search complete.")
        self.search_end_flg = True

    def url_scrap(self, area, store_junle):
        #MAX_RETRY = 3
        driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        print("starting ChromeDriver.exe....")
        wait = WebDriverWait(driver, 180)#Max wait time(second):180s
        driver.get("https://beauty.hotpepper.jp/top/")  # top page
        sr_class = driver.find_element_by_link_text(store_junle)#ジャンル選択
        sr_class.click()
        wait.until(EC.visibility_of_element_located((By.ID, "freeWordSearch1")))
        search = driver.find_element_by_id('freeWordSearch1')
        search.send_keys(area + Keys.ENTER)
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, "p.pa.bottom0.right0")))
        result_pages = driver.find_element_by_css_selector('p.pa.bottom0.right0').text
        page_num = re.split('[/ ]', result_pages)
        pages = re.sub(r"\D", "", page_num[1])
        print("pages : " + pages)
        self.page_count = int(pages)
        for i in range(int(pages)):
            if i % 100 == 0 and i > 0:
                cur_url = driver.current_url
                self.book.save(self.path)
                self.driver.quit()
                time.sleep(5)
                driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
                driver.get(cur_url)
                
            #sig.OneLineProgressMeter("掲載URLの抽出中...", self.counter, int(pages))
            try:
                html = driver.page_source
                soup = bs(html, 'lxml')
                if store_junle == 'ヘアサロン':
                    links_list = soup.select("#mainContents > ul > li > div.slnCassetteHeader > h3 > a")
                else:
                    links_list = soup.select("#mainContents > ul > li > div.slcHeadWrap > div > div.slcHeadContentsInner > h3 > a")
                print(links_list)
                for a in links_list:
                    url = a.get('href')
                    print(url)
                    r = self.sheet.max_row
                    self.sheet.cell(row=r+1, column=1, value=store_junle)#ジャンル
                    self.sheet.cell(row=r+1, column=6, value=area)#エリア
                    self.sheet.cell(row=r+1, column=8, value=url)#URL
                    #print(self.sheet.cell(row=r+1, column=8).value)
                try:
                    pre_url = driver.current_url
                    #pre_index = i
                    next_btn = driver.find_element_by_link_text("次へ")
                    next_btn.click()
                    wait.until(EC.visibility_of_all_elements_located)
                except NoSuchElementException:
                    break
            except (WebDriverException, TimeoutException):
                self.book.save(self.path)
                self.driver.quit()
                time.sleep(10)
                print("starting ChromeDriver.exe....")
                driver = webdriver.Chrome(
                    executable_path=self.driver_path, options=self.options)
                driver.get(pre_url)
                next_btn = self.driver.find_element_by_link_text("次へ")
                next_btn.click()
                wait.until(EC.visibility_of_all_elements_located)
                continue
            else:
                pass
            self.counter += 1

        self.book.save(self.path)
        print("search complete")
        driver.quit()
        #self.driver.close()

    def info_scrap(self, index):
        #conunter = 0
        wait = WebDriverWait(self.driver, 180)
        try:
            url = self.sheet.cell(row=index, column=8).value
            self.driver.get(url)
        except (WebDriverException, TimeoutException):
            self.book.save(self.path)
            self.driver.delete_all_cookies()
            self.driver.quit()
            time.sleep(30)#強制30秒待機
            self.driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
            wait = WebDriverWait(self.driver, 180)
            self.driver.get(url)
        else:
            pass
        wait.until(EC.visibility_of_all_elements_located)
        html = self.driver.page_source
        soup = bs(html, 'lxml')
        table_value = soup.select(
            'div.mT30 > table > tbody > tr > td')
        table_menu = soup.select(
            'div.mT30 > table > tbody > tr > th')
          # 住所の抽出（少々処理があるため別で書き出す）
        pref_tf = True
        prefecture = ""
        for j, e in enumerate(table_menu):
            if e.get_text() == "住所":
                all_address = table_value[j].get_text()
                prefecture_search = re.search(
                    '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)
                address_low = re.split(
                    '東京都|北海道|(?:京都|大阪)府|.{2,3}県', all_address)  # 県名とそれ以降を分離
                prefecture = prefecture_search.group()  # 県名
                if prefecture != self.sheet.cell(row=index, column=6).value:
                    self.sheet.cell(row=index, column=8, value="")
                    self.sheet.cell(row=index, column=6, value="")
                    self.sheet.cell(row=index, column=1, value="")
                    pref_tf = False
                jis_code = self.call_jis_code(prefecture)
                municipality = address_low[1]  # それ以降
                break
        # 指定エリアでないとき、下記処理を行わない
        try:
            if pref_tf:
                store_name_tag = soup.select_one('p.detailTitle > a')
                store_name = store_name_tag.get_text()
                print("店名：" + store_name)
                st_name_kana_tag = soup.select_one(
                    'div > p.fs10.fgGray')
                st_name_kana = st_name_kana_tag.get_text()
                print("店名カナ：" + st_name_kana)
                try:
                    tel_tag = soup.select_one(
                        'div.mT30 > table > tbody > tr > td > a')
                    tel_url = tel_tag.get('href')
                    respons_tel = rq.get(tel_url)
                    html_tel = respons_tel.text
                    soup_tel = bs(html_tel, 'lxml')
                    tel_num_tag = soup_tel.select_one(
                        'table > tr > td')
                    tel_num = tel_num_tag.get_text()
                    tel_num = str(tel_num)
                    tel_num = tel_num.replace(' ', "")
                    print("TEL : " + tel_num)
                except:
                    tel_num = ""
                    pass
                    # ヘッダー画像の有無
                try:
                    self.driver.find_element_by_css_selector(
                        'div.slnHeaderSliderPhoto.jscViewerPhoto')
                    head_img_yn = "有"
                    # self.sheet.cell(row=index, column=14, value="有")
                except NoSuchElementException:
                    head_img_yn = "無"
                    # self.sheet.cell(row=index, column=14, value="無")
                    pass
                catch_copy_tag = soup.select_one('div > p > b > strong')
                catch_copy = catch_copy_tag.get_text()

                pankuzu_tag = soup.select('#preContents > ol > li')
                pankuzu = ""
                for pan in pankuzu_tag:
                    pankuzu += pan.get_text()
                print(pankuzu)

                slide_img_tag = soup.select(
                    'div.slnTopImgCarouselWrap.jscThumbWrap > ul > li')
                slide_cnt = len(slide_img_tag)

                # write Excel
                try:
                    self.sheet.cell(row=index, column=2, value=store_name)
                    self.sheet.cell(row=index, column=3, value=st_name_kana)
                    self.sheet.cell(row=index, column=4, value=tel_num)
                    self.sheet.cell(row=index, column=5, value=jis_code)
                    self.sheet.cell(row=index, column=6, value=prefecture)
                    self.sheet.cell(row=index, column=7, value=municipality)
                    self.sheet.cell(row=index, column=9, value=self.scrap_day())
                    self.sheet.cell(row=index, column=13, value=pankuzu)
                    self.sheet.cell(row=index, column=16, value=slide_cnt)
                    self.sheet.cell(row=index, column=17, value=catch_copy)
                    self.sheet.cell(row=index, column=14, value=head_img_yn)
                except:
                    pass
                # 他の情報処理
                try:
                    for j in range(2, len(table_value)):
                        for c in range(1, self.sheet.max_column):
                            if table_menu[j].get_text() == self.sheet.cell(row=1, column=c).value:
                                self.sheet.cell(
                                    row=index, column=c, value=table_value[j].get_text())
                                break
                except RuntimeError:
                    pass
            else:
                print("prefname is False")
        except:
            #self.book_save()
            pass

    def restart(self, index):
        self.book.save(self.path)
        self.driver.delete_all_cookies()
        self.driver.quit()
        time.sleep(5)
        self.driver = webdriver.Chrome(executable_path=self.driver_path, options=self.options)
        url = self.sheet.cell(row=index, column=8).value
        self.driver.get(url)

    def call_jis_code(self, key):
        pref_jiscode = {
            "北海道": '01',
            "青森県": '02',
            "岩手県": '03',
            "宮城県": '04',
            "秋田県": '05',
            "山形県": '06',
            "福島県": '07',
            "茨城県": '08',
            "栃木県": '09',
            "群馬県": 10,
            "埼玉県": 11,
            "千葉県": 12,
            "東京都": 13,
            "神奈川県": 14,
            "新潟県": 15,
            "富山県": 16,
            "石川県": 17,
            "福井県": 18,
            "山梨県": 19,
            "長野県": 20,
            "岐阜県": 21,
            "静岡県": 22,
            "愛知県": 23,
            "三重県": 24,
            "滋賀県": 25,
            "京都府": 26,
            "大阪府": 27,
            "兵庫県": 28,
            "奈良県": 29,
            "和歌山県": 30,
            "鳥取県": 31,
            "島根県": 32,
            "岡山県": 33,
            "広島県": 34,
            "山口県": 35,
            "徳島県": 36,
            "香川県": 37,
            "愛媛県": 38,
            "高知県": 39,
            "福岡県": 40,
            "佐賀県": 41,
            "長崎県": 42,
            "熊本県": 43,
            "大分県": 44,
            "宮崎県": 45,
            "鹿児島県": 46,
            "沖縄県": 47
        }
        code = pref_jiscode[key]
        print(code)
        return code

    def scrap_day(self):
        dt_now = datetime.datetime.now()
        year = str(dt_now.year)
        month = str(dt_now.month)
        day = str(dt_now.day)
        hour = str(dt_now.hour)
        min = str(dt_now.minute)
        data_day = year + "," + month + day + "," + hour + min
        print(data_day)
        return data_day


def resource_path(relative_path):#バイナリフィルのパスを提供
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)

def apper_adjst(path):#空白行を削除する
    book = px.load_workbook(path)
    sheet = book.worksheets[0]
    for r in range(2, sheet.max_row+1):
        print("check" + str(r))
        #sig.OneLineProgressMeter('内容チェック中...', r-1, sheet.max_row-1, args="データの内容をチェック中...です。")
        if sheet.cell(row=r, column=1).value == "" or sheet.cell(row=r, column=1).value == None:
            sheet.delete_rows(r)
    book.save(path)

def check(path):
    book = px.load_workbook(path)
    sheet = book.worksheets[0]
    for r in range(2, sheet.max_row+1):
        print("check" + str(r))
        #sig.OneLineProgressMeter('内容チェック中...', r-1, sheet.max_row-1)
        if sheet.cell(row=r, column=2).value in ("", " ", None):
            fill =  PatternFill(patternType='solid', fgColor='ffff00')
            sheet['B'+str(r)].fill = fill
            sheet.cell(row=r, column=2, value='抽出不可')
        if sheet.cell(row=r, column=1).value == "" or sheet.cell(row=r, column=1).value == None:
            return False
    book.save(path)
    return True


class Test():

    def __init__(self):
        self.job = Scraping('./kouti_test.xlsx')
        self.job.init_work_book()
        self.url_flg = False
        self.info_flg = False
        #cnt:int = 0

    def main(self):
        th1 = th.Thread(target=self.job.url_scrap, args=('高知県', 'ヘアサロン')) 
        th1.start()
        #search_process.get()
        self.url_flg = True
        # info scraiping
        #p.join()s
        complete_row = 2 #初期値2行目
        self.info_flg = True
        while self.info_flg:
            try:
                ready_row = self.job.sheet.max_row #現在の最大読み込み行数
                print("redy:"+str(ready_row))
            except RuntimeError:
                print(ready_row)
                ready_row = self.job.sheet.max_row #現在の最大読み込み行数
                pass
            #print(ready_row)
            
            if ready_row != 1 and th1.is_alive != True and self.url_flg == True: 
                #url_scrap終了時
                print("url search end")
                self.url_flg = False
            
            if self.url_flg == False and complete_row == self.job.sheet.max_row and complete_row != 1:
                self.info_flg = False
                print("break")
                break
            
            #print("compleate row / loaded row : " + str(complete_row) + "/" + str(ready_row))
            for row in range(complete_row, ready_row+1):
                if (self.job.sheet.cell(row=row, column=8).value != None #URL抽出済
                    and self.job.sheet.cell(row=row, column=2).value == None):#info_scrap未実施                     
                    print("scrap:" + str(row))
                    self.job.info_scrap(row)
                else:#URL未抽出行に到達
                    complete_row = row #次回開始行の更新
                    break #更新のためbreak
            
    
        self.job.driver.quit()
        self.job.book.save(self.job.path)

if __name__ == "__main__":

    test = Test()
    test.main()
    #test.main()

